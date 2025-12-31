import openpyxl
import os
import datetime

print("Working directory",os.getcwd())
file_path = r"C:\Users\rucku\OneDrive\Desktop\Ruckmani\Others\Python\FirstExcel.xlsx"

wb=openpyxl.load_workbook(file_path)
print(type(wb))

ws=wb.active
ws["D3"]="=SUM(A3:C3)"

mydict={
"Name": "ruckmani",
"Flu shot": "Yes",
"Date": datetime.datetime.now().strftime("%Y-%m-%d")}

print("My dictionary:", mydict.values())

for row, (key,value) in enumerate(mydict.items(), start=10):
	ws.cell(row, column=1, value=key)
	ws.cell(row, column=2, value=value)


wb.save(file_path)

print("Excel data is read")